#!/usr/bin/env python3
""" The Transaction Statement Excel generator """

import pandas as pd
from models.user import User
from os import getenv
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import numbers, Font
from typing import Dict, Literal


currency = getenv('DEFAULT_CURRENCY', '$')


class Statement:
    """ The statement class """

    @classmethod
    def prepare_statement(self, user: User, period:Dict[str, str] = None, format: Literal['xlsx', 'pdf'] = 'xlsx') -> None:
        """ Returns the HTML Excel workbook """
        txns, summary = user.generate_statement(period)

        df = pd.DataFrame({
            'Date of Transaction': [d['Date_occurred'] for d in txns],
            'Description': [des['Description'] for des in txns],
            'Transaction Type': [t['Type'] for t in txns],
            'Amount': [a['Amount'] for a in txns],
            'Balance': [b['Balance'] for b in txns]
        })

        # Setup the intro data
        summary_df = pd.DataFrame({
            'Description': ['User Name', 'Email', 'From', 'To', 'Credits', 'Debits',
                'Opening Balance', 'Closing Balance', 'Balance'],
            'Result': [f'{user.first_name} {user.last_name}', user.email,
                period['from'], period['to'], summary['Total_credit'],
                summary['Total_debit'], summary['Opening_balance'],
                summary['Closing_balance'], summary['Balance']]
            })

        df['Date of Transaction'] = pd.to_datetime(df['Date of Transaction'])

        with pd.ExcelWriter('statement.xlsx', engine='openpyxl', mode='w') as writer:
            # Write dataframes to Excel first
            summary_df.to_excel(writer, index=False, sheet_name='Statement', startrow=7)
            df.to_excel(writer, index=False, sheet_name='Statement', startrow=18)

            # Get the actual worksheet created by pandas
            worksheet = writer.sheets['Statement']

            # Add a logo
            logo = Image('static/images/logo.png')
            logo.width *= 0.5
            logo.height *= 0.3
            logo.anchor = 'A1'
            worksheet.add_image(logo)

            # Conditional formatting for negative balances
            values_col_idx = df.columns.get_loc('Balance') + 1  # 1-based
            start_row = 18
            end_row = start_row + len(df) - 1
            column_letter = get_column_letter(values_col_idx)
            cell_range = f'{column_letter}{start_row}:{column_letter}{end_row}'

            red_font = Font(color='9C0006')
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='lessThan', formula=['0'], font=red_font)
            )

            # Apply currency formatting
            for row in range(start_row, end_row + 1):
                cell = worksheet.cell(row=row, column=values_col_idx)
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            # Auto-adjust column widths
            for idx, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)
                max_length = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[col_letter].width = max_length + 2
